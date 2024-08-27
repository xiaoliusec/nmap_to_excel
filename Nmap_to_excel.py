import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def center_window(window, screen_width, screen_height):
    window_width = 700
    window_height = 400
    x_coor = (screen_width / 2) - (window_width / 2)
    y_coor = (screen_height / 2) - (window_height / 2)
    window.geometry(f'{window_width}x{window_height}+{int(x_coor)}+{int(y_coor)}')

def clear_log():
    log_text.config(state=tk.NORMAL)
    log_text.delete(1.0, tk.END)
    log_text.config(state=tk.DISABLED)

def browse_file_or_folder():
    selected = selection_var.get()
    if selected == 'file':
        filepath = filedialog.askopenfilename(filetypes=[("XML files", "*.xml")])
        if filepath:
            input_path_var.set(filepath)
    elif selected == 'folder':
        folderpath = filedialog.askdirectory()
        if folderpath:
            input_path_var.set(folderpath)

def convert_to_excel():
    input_path = input_path_var.get()
    if input_path:
        if os.path.isfile(input_path):  # 如果输入的是单个文件
            try:
                progress_bar['maximum'] = 1
                progress_bar['value'] = 0
                root.update_idletasks()
                
                host_data = parse_nmap_xml(input_path)
                excel_file = os.path.splitext(input_path)[0] + '.xlsx'
                write_to_excel(host_data, excel_file)

                log_text.config(state=tk.NORMAL)
                log_text.insert(tk.END, f"{input_path} 中的数据已写入 {excel_file}\n")
                log_text.config(state=tk.DISABLED) 
                
                progress_bar['value'] = 1
                root.update_idletasks()
            except Exception as e:
                log_text.config(state=tk.NORMAL)
                log_text.insert(tk.END, f"处理 {input_path} 时出错: {e}\n")
                log_text.config(state=tk.DISABLED)
        elif os.path.isdir(input_path):  # 如果输入的是目录
            xml_files = [f for f in os.listdir(input_path) if f.endswith('.xml')]
            progress_bar['maximum'] = len(xml_files)
            progress_bar['value'] = 0
            log_text.delete(1.0, tk.END)
            for filename in xml_files:
                xml_file = os.path.join(input_path, filename)
                try:
                    host_data = parse_nmap_xml(xml_file)
                    excel_file = os.path.splitext(xml_file)[0] + '.xlsx'
                    write_to_excel(host_data, excel_file)
                    log_text.config(state=tk.NORMAL)
                    log_text.insert(tk.END, f"{xml_file} 中的数据已写入 {excel_file}\n")
                    log_text.config(state=tk.DISABLED)
                    progress_bar['value'] += 1
                    root.update_idletasks()
                except Exception as e:
                    log_text.config(state=tk.NORMAL)
                    log_text.insert(tk.END, f"处理 {xml_file} 时出错: {e}\n")
                    log_text.config(state=tk.DISABLED)
                    progress_bar['value'] += 1
                    root.update_idletasks()

def parse_nmap_xml(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    host_data = []

    for host in root.findall('host'):
        address = host.find('address').get('addr')
        status = host.find('status').get('state')

        ports = []
        for port in host.find('ports').findall('port'):
            port_id = port.get('portid')
            protocol = port.get('protocol')
            state = port.find('state').get('state')
            service = port.find('service')
            service_name = service.get('name') if service is not None else 'unknown'
            ports.append((port_id, protocol, state, service_name))


        os_info = []
        os_element = host.find('os')
        if os_element is not None:
            best_os_match = None
            max_accuracy = 0
            for os_match in os_element.findall('osmatch'):
                accuracy = int(os_match.get('accuracy'))
                if accuracy > max_accuracy:
                    max_accuracy = accuracy
                    best_os_match = os_match.get('name')
            if best_os_match:
                os_info.append((best_os_match, max_accuracy))

        host_data.append({
            'address': address,
            'status': status,
            'ports': ports,
            'os_info': os_info
        })

    return host_data

def write_to_excel(host_data, excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Nmap探测结果"

    headers = ['主机IP', '存活状态', '端口', '协议', '端口状态', '服务', '操作系统类型', '操作系统识别准确率']
    for col_num, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_num)
        c.value = header

    header_font = Font(color="FFFFFF", bold=True)
    header_fill = PatternFill(start_color="80807F", end_color="80807F", fill_type="solid")
    header_alignment = Alignment(horizontal="left")

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 写入数据
    row_num = 2
    for host in host_data:
        if host['os_info']:
            os_name, os_accuracy = host['os_info'][0]
            for port_info in host['ports']:
                ws.cell(row=row_num, column=1, value=host['address'])
                ws.cell(row=row_num, column=2, value=host['status'])
                ws.cell(row=row_num, column=3, value=port_info[0])
                ws.cell(row=row_num, column=4, value=port_info[1])
                ws.cell(row=row_num, column=5, value=port_info[2])
                ws.cell(row=row_num, column=6, value=port_info[3])
                ws.cell(row=row_num, column=7, value=os_name)
                ws.cell(row=row_num, column=8, value=os_accuracy)
                row_num += 1

    merge_identical_cells(ws)

    
    for col in range(1, 9):
        if col <= 6:
            ws.column_dimensions[get_column_letter(col)].width = 15
        else:
            ws.column_dimensions[get_column_letter(col)].width = 20

    # 保存
    wb.save(excel_file)

def merge_identical_cells(ws):
    merged_ranges = set()
    prev_ip = None
    start_row = None
    for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        current_ip = row[0]
        if current_ip != prev_ip:
            if start_row is not None:
                end_row = idx - 1
                if start_row <= end_row:
                    merged_ranges.add(f"A{start_row}:A{end_row}")
            start_row = idx
        prev_ip = current_ip

    if start_row is not None:
        end_row = ws.max_row
        if start_row <= end_row:
            merged_ranges.add(f"A{start_row}:A{end_row}")

    for range_str in merged_ranges:
        ws.merge_cells(range_str)

if __name__ == "__main__":
    global root
    root = tk.Tk()
    root.title("Nmap探测结果转换为Excel表格")

    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    center_window(root, screen_width, screen_height)

    icon_path = 'icon1.ico'
    root.wm_iconbitmap(icon_path)

    input_path_var = tk.StringVar()
    selection_var = tk.StringVar(value='file')  # 默认选择单个文件

    frame = ttk.Frame(root, padding="10")
    frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    ttk.Radiobutton(frame, text="单个文件", variable=selection_var, value='file').grid(column=0, row=0, sticky=tk.W)
    ttk.Radiobutton(frame, text="文件夹", variable=selection_var, value='folder').grid(column=1, row=0, sticky=tk.W)

    ttk.Label(frame, text="选择XML文件或文件夹:").grid(column=0, row=1, sticky=tk.W)
    ttk.Entry(frame, width=50, textvariable=input_path_var).grid(column=0, row=2, sticky=(tk.W, tk.E))
    ttk.Button(frame, text="浏览...", command=browse_file_or_folder).grid(column=1, row=2)

    ttk.Label(frame, text="转换进度:").grid(column=0, row=3, sticky=tk.W)
    progress_bar = ttk.Progressbar(frame, orient="horizontal", length=200, mode="determinate")
    progress_bar.grid(column=0, row=4, columnspan=2, sticky=(tk.W, tk.E))

    ttk.Button(frame, text="转换为Excel", command=convert_to_excel).grid(column=0, row=5, columnspan=2)

    ttk.Button(frame, text="清空日志", command=clear_log).grid(column=1, row=5, columnspan=2)

    log_text = tk.Text(frame, height=15, width=100)
    log_text.config(font=("Consolas", 10))
    log_text.grid(column=0, row=6, columnspan=2)

    root.columnconfigure(0, weight=1)
    root.rowconfigure(0, weight=1)
    frame.columnconfigure(0, weight=1)
    frame.rowconfigure(6, weight=1)

    root.mainloop()
